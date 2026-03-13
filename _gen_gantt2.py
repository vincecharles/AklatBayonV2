"""
AklatBayon v2 — Full Project Gantt Chart (Feb 2026 - Mar 2027)
Monthly columns, covering prototype through Laravel transition.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart"

ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_view.showGridLines = False

# Colors
WHITE = 'FFFFFF'
DARK_BG = '1E3A5F'
LIGHT_GRAY = 'F5F5F5'
BORDER_GRAY = 'D9D9D9'
TODAY_BG = 'E8F4FD'

PHASE_COLORS = {
    1: '2E86AB', 2: '564787', 3: 'E07A5F', 4: '3D9970', 5: 'D4A017',
    6: '7B68EE', 7: '2196F3', 8: 'E74C3C', 9: 'FF9800', 10: '00BCD4',
    11: '8E44AD', 12: '2ECC71', 13: 'E67E22', 14: '1ABC9C', 15: 'C0392B',
}

BAR_COLORS = {
    'done': '4CAF50', 'progress': '2196F3',
    'pending': 'FFB74D', 'milestone': 'E91E63',
}

STATUS_LABELS = {
    'done':      ('Complete',    'D4EDDA', '155724'),
    'progress':  ('In Progress', 'CCE5FF', '004085'),
    'pending':   ('Not Started', 'FFF3CD', '856404'),
    'milestone': ('Milestone',   'FCE4EC', 'C2185B'),
}

thin_border = Border(
    left=Side(style='thin', color=BORDER_GRAY),
    right=Side(style='thin', color=BORDER_GRAY),
    top=Side(style='thin', color=BORDER_GRAY),
    bottom=Side(style='thin', color=BORDER_GRAY),
)

# Timeline: Feb 2026 to Mar 2027 = 14 months
months = []
for y in [2026, 2027]:
    start_m = 2 if y == 2026 else 1
    end_m = 12 if y == 2026 else 3
    for m in range(start_m, end_m + 1):
        months.append((y, m))

MONTH_NAMES = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

TODAY_MONTH = (2026, 3)

# Task data: (name, start_month_idx, end_month_idx, status)
# month index 0 = Feb 2026, 1 = Mar 2026, etc.
def mi(y, m):
    """Month index from year,month."""
    for i, (yy, mm) in enumerate(months):
        if yy == y and mm == m:
            return i
    return 0

phases = [
    (1, 'Project Setup & Foundation', [
        ('Project init, repo setup, Netlify config',    mi(2026,2), mi(2026,2), 'done'),
        ('CSS styling & responsive design',              mi(2026,2), mi(2026,2), 'done'),
        ('Landing page, login, welcome pages',           mi(2026,2), mi(2026,2), 'done'),
        ('Package.json & dependency management',         mi(2026,3), mi(2026,3), 'done'),
    ]),
    (2, 'Core UI & Authentication (Prototype)', [
        ('Auth system, session management',              mi(2026,2), mi(2026,2), 'done'),
        ('localStorage data store (store.js)',           mi(2026,2), mi(2026,2), 'done'),
        ('Permission-based sidebar & app shell',         mi(2026,2), mi(2026,2), 'done'),
        ('Dashboard with stats & charts',                mi(2026,2), mi(2026,3), 'done'),
        ('Dark mode, profanity filter',                  mi(2026,2), mi(2026,2), 'done'),
    ]),
    (3, 'User, Role & Student Management', [
        ('User CRUD & role management (7 roles)',        mi(2026,2), mi(2026,2), 'done'),
        ('23 permissions across 8 groups',               mi(2026,2), mi(2026,2), 'done'),
        ('Student records CRUD',                         mi(2026,2), mi(2026,2), 'done'),
    ]),
    (4, 'Catalog Management', [
        ('Books, Authors, Publishers, Categories CRUD',  mi(2026,2), mi(2026,2), 'done'),
        ('LCC browser & LOC API integration',            mi(2026,2), mi(2026,2), 'done'),
        ('Barcode generation & advanced OPAC',           mi(2026,3), mi(2026,3), 'done'),
    ]),
    (5, 'Circulation, Reservations & Fines', [
        ('Book issue, return & renewal',                 mi(2026,2), mi(2026,3), 'done'),
        ('Reservation system (FIFO queue, expiry)',      mi(2026,3), mi(2026,3), 'done'),
        ('Auto fine computation & block threshold',      mi(2026,3), mi(2026,3), 'done'),
        ('Borrower profile with QR code',                mi(2026,3), mi(2026,3), 'done'),
    ]),
    (6, 'Administration & Reporting', [
        ('Fine management & My Fines page',              mi(2026,2), mi(2026,3), 'done'),
        ('Inventory, reports & analytics',               mi(2026,2), mi(2026,2), 'done'),
        ('RFID attendance tracking',                     mi(2026,2), mi(2026,2), 'done'),
        ('Audit logs & system settings (30+ params)',    mi(2026,2), mi(2026,3), 'done'),
    ]),
    (7, 'Serverless Backend (Neon PostgreSQL)', [
        ('Neon DB provisioning & Drizzle schema (16 tables)', mi(2026,3), mi(2026,3), 'done'),
        ('REST API serverless function (api.mts)',       mi(2026,3), mi(2026,3), 'done'),
        ('Database seed function (seed.mts)',            mi(2026,3), mi(2026,3), 'done'),
        ('Frontend API client & hybrid store',           mi(2026,3), mi(2026,3), 'done'),
        ('Deploy & seed production database',            mi(2026,3), mi(2026,3), 'progress'),
    ]),
    (8, 'Prototype Security & QA', [
        ('Password hashing & JWT tokens',                mi(2026,3), mi(2026,4), 'pending'),
        ('Server-side auth middleware & rate limiting',   mi(2026,4), mi(2026,4), 'pending'),
        ('Input sanitization & RBAC on API',             mi(2026,4), mi(2026,4), 'pending'),
        ('API testing & data integrity validation',      mi(2026,4), mi(2026,5), 'pending'),
        ('User acceptance testing (UAT)',                 mi(2026,5), mi(2026,5), 'pending'),
    ]),
    (9, 'Prototype Launch & Documentation', [
        ('Final prototype deploy & smoke test',          mi(2026,5), mi(2026,5), 'pending'),
        ('Capstone Chapter 1-5 documentation',           mi(2026,3), mi(2026,5), 'progress'),
        ('Capstone defense (Prototype)',                 mi(2026,5), mi(2026,5), 'milestone'),
    ]),
    (10, 'Laravel Backend Setup', [
        ('Laravel project scaffolding',                  mi(2026,6), mi(2026,6), 'pending'),
        ('Database migrations (16 tables to Eloquent)',  mi(2026,6), mi(2026,7), 'pending'),
        ('Eloquent models & relationships',              mi(2026,7), mi(2026,7), 'pending'),
        ('Auth scaffolding (Breeze/Fortify)',             mi(2026,7), mi(2026,7), 'pending'),
        ('Spatie permission package (7 roles, 23 perms)',mi(2026,7), mi(2026,8), 'pending'),
        ('Laravel API routes & controllers',             mi(2026,8), mi(2026,9), 'pending'),
    ]),
    (11, 'Vue.js + Inertia Frontend', [
        ('Vue 3 + Inertia.js setup with Laravel',       mi(2026,8), mi(2026,8), 'pending'),
        ('Bootstrap integration & layout components',    mi(2026,8), mi(2026,9), 'pending'),
        ('Port pages to Vue components (23 pages)',      mi(2026,9), mi(2026,11), 'pending'),
        ('Dashboard, sidebar & navigation in Vue',       mi(2026,9), mi(2026,10), 'pending'),
        ('Reactive forms & real-time validation',        mi(2026,10), mi(2026,11), 'pending'),
    ]),
    (12, 'Full-Stack Feature Migration', [
        ('Catalog module (CRUD, LOC API, LCC, barcode)', mi(2026,9), mi(2026,10), 'pending'),
        ('Circulation module (issue, return, renew)',     mi(2026,10), mi(2026,11), 'pending'),
        ('Reservations, fines & auto-computation',       mi(2026,10), mi(2026,11), 'pending'),
        ('RFID attendance & audit logs',                 mi(2026,11), mi(2026,11), 'pending'),
        ('Reports, inventory & settings',                mi(2026,11), mi(2026,12), 'pending'),
    ]),
    (13, 'Production Security & Hardening', [
        ('Bcrypt password hashing in Laravel',           mi(2026,12), mi(2026,12), 'pending'),
        ('Sanctum/Passport API authentication',          mi(2026,12), mi(2026,12), 'pending'),
        ('Laravel middleware (auth, roles, throttle)',    mi(2026,12), mi(2027,1), 'pending'),
        ('CSRF, XSS & SQL injection protection',         mi(2027,1), mi(2027,1), 'pending'),
        ('Server-side input validation (Form Requests)',  mi(2027,1), mi(2027,1), 'pending'),
    ]),
    (14, 'Testing & QA (Production)', [
        ('PHPUnit & Pest feature tests',                 mi(2027,1), mi(2027,2), 'pending'),
        ('Vue component testing (Vitest)',                mi(2027,1), mi(2027,2), 'pending'),
        ('Cross-browser & mobile testing',               mi(2027,2), mi(2027,2), 'pending'),
        ('Load testing & performance optimization',      mi(2027,2), mi(2027,2), 'pending'),
        ('UAT with FEATI library staff & students',      mi(2027,2), mi(2027,3), 'pending'),
    ]),
    (15, 'Production Launch', [
        ('Production server deployment',                 mi(2027,3), mi(2027,3), 'pending'),
        ('Data migration from prototype to production',  mi(2027,3), mi(2027,3), 'pending'),
        ('Final documentation & handover',               mi(2027,3), mi(2027,3), 'pending'),
        ('GO LIVE (Production)',                         mi(2027,3), mi(2027,3), 'milestone'),
    ]),
]

# Column setup
ws.column_dimensions['A'].width = 1.5
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 11
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10

MONTH_COL_START = 6  # column F

for i in range(len(months)):
    ws.column_dimensions[get_column_letter(MONTH_COL_START + i)].width = 8.5

# Row 1: Title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
c = ws.cell(row=1, column=1)
c.value = "AklatBayon v2 \u2014 Full Project Gantt Chart"
c.font = Font(name='Calibri', size=16, bold=True, color=DARK_BG)
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 30

# Row 2: Subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
c = ws.cell(row=2, column=1)
c.value = "FEATI University Library Management System \u2014 Feb 2026 to Mar 2027 (Prototype to Production)"
c.font = Font(name='Calibri', size=9, color='666666')
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 18

# Row 3: Year headers
year_spans = {}
for i, (y, m) in enumerate(months):
    if y not in year_spans:
        year_spans[y] = [i, i]
    year_spans[y][1] = i

for y, (si, ei) in year_spans.items():
    sc = MONTH_COL_START + si
    ec = MONTH_COL_START + ei
    ws.merge_cells(start_row=3, start_column=sc, end_row=3, end_column=ec)
    c = ws.cell(row=3, column=sc)
    c.value = str(y)
    c.font = Font(name='Calibri', size=9, bold=True, color=WHITE)
    c.fill = PatternFill('solid', fgColor=DARK_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    for ci in range(sc, ec + 1):
        ws.cell(row=3, column=ci).border = thin_border

ws.row_dimensions[3].height = 18

# Row 4: Column headers
HR = 4
for val, col, ah in [('', 1, 'center'), ('Task Name', 2, 'left'),
                      ('Status', 3, 'center'), ('Start', 4, 'center'), ('End', 5, 'center')]:
    c = ws.cell(row=HR, column=col)
    c.value = val
    c.font = Font(name='Calibri', size=8, bold=True, color=WHITE)
    c.fill = PatternFill('solid', fgColor=DARK_BG)
    c.alignment = Alignment(horizontal=ah, vertical='center')
    c.border = thin_border

# Month headers
for i, (y, m) in enumerate(months):
    col = MONTH_COL_START + i
    is_today = (y, m) == TODAY_MONTH
    c = ws.cell(row=HR, column=col)
    c.value = MONTH_NAMES[m]
    c.font = Font(name='Calibri', size=8, bold=True, color=WHITE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border
    if is_today:
        c.fill = PatternFill('solid', fgColor='FF6B35')
    else:
        c.fill = PatternFill('solid', fgColor=DARK_BG)

ws.row_dimensions[HR].height = 22
ws.freeze_panes = ws.cell(row=HR + 1, column=MONTH_COL_START)

# Data rows
current_row = HR + 1

def month_label(idx):
    y, m = months[idx]
    return f"{MONTH_NAMES[m]} {y}"

for phase_num, phase_name, tasks in phases:
    pc = PHASE_COLORS.get(phase_num, '999999')

    # Phase row
    ws.row_dimensions[current_row].height = 24
    ws.cell(row=current_row, column=1).fill = PatternFill('solid', fgColor=pc)
    ws.cell(row=current_row, column=1).border = thin_border

    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
    c = ws.cell(row=current_row, column=2)
    c.value = f"Phase {phase_num}: {phase_name}"
    c.font = Font(name='Calibri', size=9, bold=True, color=pc)
    c.fill = PatternFill('solid', fgColor='F8F9FA')
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = thin_border
    for ci in range(3, 6):
        ws.cell(row=current_row, column=ci).fill = PatternFill('solid', fgColor='F8F9FA')
        ws.cell(row=current_row, column=ci).border = thin_border

    # Phase span bar
    ps = min(t[1] for t in tasks)
    pe = max(t[2] for t in tasks)
    for i in range(len(months)):
        col = MONTH_COL_START + i
        cell = ws.cell(row=current_row, column=col)
        cell.border = thin_border
        if ps <= i <= pe:
            cell.fill = PatternFill('solid', fgColor=pc)
        else:
            cell.fill = PatternFill('solid', fgColor='F8F9FA')

    current_row += 1

    # Tasks
    for task_name, start_i, end_i, status in tasks:
        is_even = current_row % 2 == 0
        row_bg = LIGHT_GRAY if is_even else WHITE

        ws.row_dimensions[current_row].height = 20

        # Phase strip
        ws.cell(row=current_row, column=1).fill = PatternFill('solid', fgColor=pc)
        ws.cell(row=current_row, column=1).border = thin_border

        # Task name
        c = ws.cell(row=current_row, column=2)
        c.value = task_name
        c.font = Font(name='Calibri', size=8, color='333333')
        c.fill = PatternFill('solid', fgColor=row_bg)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = thin_border

        # Status
        sl, sb, sc_c = STATUS_LABELS[status]
        c = ws.cell(row=current_row, column=3)
        c.value = sl
        c.font = Font(name='Calibri', size=7, bold=True, color=sc_c)
        c.fill = PatternFill('solid', fgColor=sb)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

        # Start
        c = ws.cell(row=current_row, column=4)
        c.value = month_label(start_i)
        c.font = Font(name='Calibri', size=7, color='666666')
        c.fill = PatternFill('solid', fgColor=row_bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

        # End
        c = ws.cell(row=current_row, column=5)
        c.value = month_label(end_i)
        c.font = Font(name='Calibri', size=7, color='666666')
        c.fill = PatternFill('solid', fgColor=row_bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

        # Month bars
        bc = BAR_COLORS[status]
        for i in range(len(months)):
            col = MONTH_COL_START + i
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            is_today_m = months[i] == TODAY_MONTH

            if start_i <= i <= end_i:
                cell.fill = PatternFill('solid', fgColor=bc)
            elif is_today_m:
                cell.fill = PatternFill('solid', fgColor=TODAY_BG)
            else:
                cell.fill = PatternFill('solid', fgColor=row_bg)

        current_row += 1

# Spacer
current_row += 1

# Legend
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
c = ws.cell(row=current_row, column=2)
c.value = "Legend:   \u25a0 Complete    \u25a0 In Progress    \u25a0 Not Started    \u25c6 Milestone    \u258c Current Month (Mar 2026)"
c.font = Font(name='Calibri', size=8, color='666666')
c.alignment = Alignment(horizontal='left', vertical='center')

# Stack info
current_row += 1
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
c = ws.cell(row=current_row, column=2)
c.value = "Phases 1-9: Prototype (Vanilla JS + Netlify + Neon PostgreSQL)    |    Phases 10-15: Production (Vue + Inertia + Laravel + Bootstrap)"
c.font = Font(name='Calibri', size=8, bold=True, color='1E3A5F')
c.alignment = Alignment(horizontal='left', vertical='center')

# Footer
current_row += 1
ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=12)
c = ws.cell(row=current_row, column=2)
c.value = "AklatBayon v2 \u2014 FEATI University \u2014 IT415 Capstone \u2014 Generated March 11, 2026"
c.font = Font(name='Calibri', size=7, italic=True, color='999999')
c.alignment = Alignment(horizontal='left', vertical='center')

# Save
output = r'c:\Users\Bins\AklatBayonV2\AklatBayon_GanttChart_v2.xlsx'
wb.save(output)
print(f'Saved: {output}')
